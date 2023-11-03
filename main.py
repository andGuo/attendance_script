from typing import Dict, List, NamedTuple

import os
import errno
from openpyxl import load_workbook

# Make sure the files starts with these strings in the same directory as this script.
TUTORIAL_LIST_FILENAME = "tutorials_merged_20230928"  # should be .xlsx
ATTENDANCE_NAMES_FILE = "bot_input"  # any text file format
OVERWRITE_MODE = True  # if false, will first reset the sheet's score to 0 before updating attendance


# These needs to be updated every week
TUTORIAL_NUMBER = 6  # the tutorial number for the week (int)
MAX_SCORE = 2  # maximum score (int) a student can get for the tutorial


def main():
    tut_list_path = find_file_path(TUTORIAL_LIST_FILENAME)
    print(f"Found file: {TUTORIAL_LIST_FILENAME} at {tut_list_path}!")
    bot_attendance_path = find_file_path(ATTENDANCE_NAMES_FILE)
    print(f"Found file: {ATTENDANCE_NAMES_FILE} at {bot_attendance_path}!")

    if not OVERWRITE_MODE:
        print(
            f"WARNING: You are about to perform an operation that will reset all existing scores in Sheet {TUTORIAL_NUMBER}."
        )
        user_input = input("Do you wish to continue (Y/n)?: ").strip().lower()

        # Check if the user's input is 'y' or 'yes'
        if user_input != "y":
            print("Aborting...")
            exit(0)
        else:
            print("Continuing...")
    else:
        print(f"Updating scores in Sheet {TUTORIAL_NUMBER} - (Tutorial {TUTORIAL_NUMBER})...")

    attendance = TakeAttendance(
        output_path=tut_list_path,
        input_path=bot_attendance_path,
        tutorial_number=TUTORIAL_NUMBER,
        overwrite_mode=OVERWRITE_MODE,
    )

    attendance.run()


def find_file_path(file_str: str) -> str:
    """
    Find the file path of the first file in the script's directory that starts with file_str

    Args:
        file_str (str): string for the file to find

    Returns:
        str: absolute path to the file
    """
    current_directory = os.path.dirname(os.path.realpath(__file__))

    for filename in os.listdir(current_directory):
        if filename.startswith(file_str):
            return os.path.join(current_directory, filename)

    raise FileNotFoundError(errno.ENOENT, os.strerror(errno.ENOENT), file_str)


class StudentAttendance(NamedTuple):
    username: str
    sid: str
    score: int | float


class TakeAttendance:
    def __init__(
        self,
        output_path: str,
        input_path: str,
        tutorial_number: int,
        overwrite_mode: bool = True,
    ) -> None:
        self.output_path: str = output_path
        self.input_path: str = input_path
        self.tutorial_number: str = f"Tutorial {tutorial_number}"
        self.overwrite_mode: bool = overwrite_mode

    def _load_student_attendance(self) -> List[str]:
        """
        Load the bot's student attendance from the file specified by self.input_path

        Returns:
            List[str]: a list of strings, each string is a student's username
        """
        try:
            with open(self.input_path, "r") as file:
                usernames = [
                    next(
                        filter(lambda x: x.startswith("#"), line.split(",", 1))
                    ).strip()
                    for line in file
                ]
            return usernames
        except Exception as e:
            print("ERROR >>> Unable to parse student attendance file!")
            raise Exception(e)

    def _load_tutorial_list(self) -> Dict[str, StudentAttendance]:
        """
        Load the tutorial student list from the file specified by self.output_path

        Returns:
            Dict[str, StudentAttendance]: a dictionary, the key is the username, the value is a StudentAttendance NamedTuple
        """
        try:
            workbook = load_workbook(
                filename=self.output_path, data_only=True, read_only=True
            )
            worksheet = workbook[self.tutorial_number]

            # Determine which columns are sid, username, and score
            for i in range(1, 4):
                col_str = worksheet.cell(1, i).value

                if col_str == "OrgDefinedId":
                    sid_col = i
                elif col_str == "Username":
                    username_col = i
                elif col_str.startswith("Tutorial"):
                    score_col = i
                else:
                    raise Exception("Undefined column name")

            tut_list_dict = {}
            num_rows = 0

            # Count number of non-empty rows
            for row in worksheet:
                if not all(col.value is None for col in row):
                    num_rows += 1

            # Create dict
            for row in range(2, num_rows + 1):
                key = worksheet.cell(row, username_col).value

                if key in tut_list_dict:
                    raise Exception("Duplicate username found")

                curr_score = worksheet.cell(row, score_col).value

                if self.overwrite_mode and curr_score is not None and curr_score > 0:
                    value = StudentAttendance(
                        username=key,
                        sid=worksheet.cell(row, sid_col).value,
                        score=curr_score if curr_score <= MAX_SCORE else MAX_SCORE,
                    )
                else:
                    value = StudentAttendance(
                        username=key,
                        sid=worksheet.cell(row, sid_col).value,
                        score=0,
                    )

                tut_list_dict[key] = value

            workbook.close()

            return tut_list_dict

        except Exception as e:
            print("ERROR >>> Unable to parse tutorial file!")
            raise Exception(e)

    def _update_attendance(
        self, usernames: List[str], tutorial_dict: Dict[str, StudentAttendance]
    ) -> List[StudentAttendance]:
        """
        Update the attendance of the students in tutorial_dict with the students in attendance

        Args:
            usernames (List[str]): list of students from the bot's attendance
            tutorial_dict (Dict[str, StudentAttendance]): dictionary of students from the tutorial list

        Returns:
            List[StudentAttendance]: list of students from the tutorial list, updated with the students in attendance
        """

        username_set = set(usernames)

        attendance_marks = {
            username: usernames.count(username) for username in username_set
        }

        for username in username_set:
            if username in tutorial_dict:
                old_score = tutorial_dict[username].score

                if attendance_marks[username] > 2:
                    raise Exception(
                        f"Invalid student attendance with username: {username} (count: {attendance_marks[username]})"
                    )

                # only use attendance score if the student does not have the max score (i.e. been checked off)
                new_score = (
                    attendance_marks[username]
                    if old_score < MAX_SCORE
                    else MAX_SCORE
                )

                tutorial_dict[username] = StudentAttendance(
                    username=username, sid=tutorial_dict[username].sid, score=new_score
                )
                print(
                    f"Updated {username:<25} {'-':<2} Old: {old_score} | New: {new_score}"
                )
            else:
                print(f"WARNING >>> Username: {username} not found in tutorial list!")

        return list(tutorial_dict.values())

    def _write_tutorial_list(self, students: List[StudentAttendance]) -> int:
        """
        Write the tutorial list to the file specified by self.output_path

        Args:
            students (List): list of StudentAttendance NamedTuples

        Returns:
            int: number of students written to file
        """
        students.sort(key=lambda x: x.sid)

        try:
            workbook = load_workbook(filename=self.output_path)
            worksheet = workbook[self.tutorial_number]

            # Determine which columns are sid, username, and score
            for i in range(1, 4):
                col_str = worksheet.cell(1, i).value

                if col_str == "OrgDefinedId":
                    sid_col = i
                elif col_str == "Username":
                    username_col = i
                elif col_str.startswith("Tutorial"):
                    score_col = i
                else:
                    raise Exception("Undefined column name")

            # Write to file
            for row, student in enumerate(students, start=2):
                worksheet.cell(row, sid_col).value = student.sid
                worksheet.cell(row, username_col).value = student.username
                worksheet.cell(row, score_col).value = (
                    student.score if student.score > 0 else None
                )

            workbook.save(filename=self.output_path)
            workbook.close()

            return row - 1

        except Exception as e:
            print("ERROR >>> Unable to write to tutorial file!")
            raise Exception(e)

    def run(self):
        student_attendance = self._load_student_attendance()
        print(f"Loaded {len(student_attendance)} students from {self.input_path}")
        attendance_dict = self._load_tutorial_list()
        print(f"Found {len(attendance_dict)} students from {self.output_path}")
        new_attendance = self._update_attendance(student_attendance, attendance_dict)
        num = self._write_tutorial_list(new_attendance)
        print(f"Successfully wrote {num} students to {self.output_path}!")


if __name__ == "__main__":
    main()
